import pandas as pd
import openpyxl


class Data:
    def __init__(self, path, sheet_name):
        print(f"Reading .xlsx file from {path} ... ")
        self.data = openpyxl.load_workbook(path)
        self.data = self.data.get_sheet_by_name(sheet_name)
        
    def GetOriginalData(self, row):
        original_data = {
            'name' : self.data.cell(row=row, column=3).value,
            'addr' : self.data.cell(row=row, column=13).value
        }
        return original_data

    def GetTranslatedData(self, row):
        translated_data = {
            'name' : self.data.cell(row=row, column=4).value,
            'addr' : self.data.cell(row=row, column=14).value,
            'addr_detail' : self.data.cell(row=row, column=19).value,
            'product' : self.data.cell(row=row, column=21).value,
            'tel' : self.data.cell(row=row, column=23).value,
            'open' : self.data.cell(row=row, column=25).value,
            'close' : self.data.cell(row=row, column=28).value,
            'site_url' : self.data.cell(row=row, column=26).value,
            'refund_counter_operator' : self.data.cell(row=row, column=31).value,
            'refund_method' : self.data.cell(row=row, column=33).value,
            'lang_info' : self.data.cell(row=row, column=35).value,
            'parking_lot' : self.data.cell(row=row, column=37).value,
            'storage' : self.data.cell(row=row, column=39).value,
            'payment_method' : self.data.cell(row=row, column=42).value,
            'toilet_kr' : self.data.cell(row=row, column=43).value,
            'toilet' : self.data.cell(row=row, column=45).value
        }
        return translated_data